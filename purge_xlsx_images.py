# -*- coding: utf-8 -*-
import json, io, concurrent.futures
import requests
from PIL import Image as PILImage
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC_XLSX = "Productos_ocultados_purga.xlsx"      # datos tabulares ya resueltos (categorías, etc.)
OUT_XLSX = "Productos_ocultados_con_imagenes.xlsx"
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36"
THUMB = 96  # px

# 1) Leer datos del Excel anterior
wb_src = openpyxl.load_workbook(SRC_XLSX)
ws_src = wb_src.active
headers = [c.value for c in ws_src[1]]
idx = {h: i for i, h in enumerate(headers)}
rows = []
for r in ws_src.iter_rows(min_row=2, values_only=True):
    rows.append(r)
print("Filas leidas:", len(rows))

# 2) Descargar + miniaturizar imagenes en paralelo
def fetch_thumb(url):
    if not url or not str(url).startswith("http"):
        return None
    try:
        resp = requests.get(str(url), headers={"User-Agent": UA}, timeout=10)
        if resp.status_code != 200 or not resp.content:
            return None
        im = PILImage.open(io.BytesIO(resp.content))
        im = im.convert("RGB")
        im.thumbnail((THUMB, THUMB))
        bio = io.BytesIO()
        im.save(bio, format="PNG")
        bio.seek(0)
        return bio
    except Exception:
        return None

img_col = idx.get("IMAGEN")
urls = [row[img_col] if img_col is not None else None for row in rows]
thumbs = [None] * len(urls)
ok = 0
with concurrent.futures.ThreadPoolExecutor(max_workers=24) as ex:
    futs = {ex.submit(fetch_thumb, u): i for i, u in enumerate(urls)}
    for fut in concurrent.futures.as_completed(futs):
        i = futs[fut]
        res = fut.result()
        thumbs[i] = res
        if res is not None:
            ok += 1
print("Imagenes descargadas OK:", ok, "de", len(urls))

# 3) Construir el nuevo Excel
OUT_COLS = [
    ("IMAGEN", None, 15),
    ("SKU", "SKU", 12),
    ("NOMBRE", "NOMBRE", 50),
    ("MARCA", "MARCA", 10),
    ("CATEGORIA", "CATEGORIA", 22),
    ("PRECIO", "PRECIO", 11),
    ("PRECIO_OFERTA", "PRECIO_OFERTA", 12),
    ("PRECIO_MAYORISTA", "PRECIO_MAYORISTA", 13),
    ("COSTO", "COSTO", 10),
    ("STOCK_ORIGINAL", "STOCK_ORIGINAL", 10),
    ("COD_BARRAS", "CODIGO_BARRAS", 16),
    ("SECCION", "SECCION", 8),
    ("ID_DOCUMENTO", "ID_DOCUMENTO", 22),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Ocultados con imagen"

hdr_fill = PatternFill("solid", fgColor="111827")
hdr_font = Font(bold=True, color="FFFFFF", size=10)
thin = Side(style="thin", color="E5E7EB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Encabezados
for ci, (label, _, width) in enumerate(OUT_COLS, start=1):
    c = ws.cell(row=1, column=ci, value=label)
    c.fill = hdr_fill; c.font = hdr_font
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions[get_column_letter(ci)].width = width
ws.row_dimensions[1].height = 22
ws.freeze_panes = "A2"

# Filas
for ri, row in enumerate(rows, start=2):
    ws.row_dimensions[ri].height = 74
    for ci, (label, src_key, _) in enumerate(OUT_COLS, start=1):
        cell = ws.cell(row=ri, column=ci)
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=(label == "NOMBRE"))
        if src_key is None:
            continue
        val = row[idx[src_key]] if src_key in idx else None
        cell.value = val
    # incrustar imagen
    th = thumbs[ri - 2]
    if th is not None:
        try:
            xi = XLImage(th)
            xi.width = THUMB; xi.height = THUMB
            ws.add_image(xi, f"A{ri}")
        except Exception:
            ws.cell(row=ri, column=1, value="(img)")
    else:
        ws.cell(row=ri, column=1, value="sin imagen").alignment = Alignment(horizontal="center", vertical="center")

wb.save(OUT_XLSX)
print("Guardado:", OUT_XLSX)
